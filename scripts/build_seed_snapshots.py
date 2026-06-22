from __future__ import annotations

import json
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SNAPSHOTS_OUT = ROOT / "src" / "data" / "seedSnapshots.json"

WORKBOOKS = [
    {
        "path": Path("/Users/drewwelker/Downloads/6-13-2026 DDT Tracker.xlsx"),
        "location": "Touhy",
        "week_start": None,
    },
    {
        "path": Path("/Users/drewwelker/Downloads/DDT Devon 6-13-26.xlsx"),
        "location": "Devon",
        "week_start": "2026-06-13",
    },
    {
        "path": Path("/Users/drewwelker/Downloads/6-20-2026 DDT DEVON WOLF.xlsx"),
        "location": "Devon",
        "week_start": "2026-06-20",
    },
]

DAY_SHEETS = {"SAT", "SUN", "MON", "TUE", "WED", "THU", "FRI"}
DAY_OFFSETS = {"SAT": 0, "SUN": 1, "MON": 2, "TUE": 3, "WED": 4, "THU": 5, "FRI": 6}


def to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.year in {1899, 1900}:
            return value.strftime("%H:%M")
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def norm(value: Any) -> str:
    return to_text(value).upper().replace("\xa0", " ").strip()


def sheet_date(ws, week_start: str | None) -> str:
    if week_start and ws.title.upper() in DAY_OFFSETS:
        return (date.fromisoformat(week_start) + timedelta(days=DAY_OFFSETS[ws.title.upper()])).isoformat()
    value = ws["C1"].value
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"{ws.title} does not have a date in C1")


def detect_shift(row_values: list[Any], current_shift: str) -> str:
    joined = " ".join(norm(value) for value in row_values)
    if "TOUR 1" in joined:
        return "Tour 1"
    if "TOUR 2" in joined:
        return "Tour 2"
    if "TOUR 3" in joined:
        return "Tour 3"
    return current_shift


def find_index(headers: list[str], predicate) -> int | None:
    for index, header in enumerate(headers):
        if predicate(header):
            return index
    return None


def build_columns(headers: list[str]) -> dict[str, Any]:
    skd_indices = [index for index, header in enumerate(headers) if "SKD KDT" in header]
    flight_indices = [index for index, header in enumerate(headers) if header == "FLT"]
    ddt_scheduled = skd_indices[0] if skd_indices else None
    kat_scheduled = skd_indices[1] if len(skd_indices) > 1 else None
    ddt_actual = find_index(
        headers[ddt_scheduled + 1 :] if ddt_scheduled is not None else [],
        lambda header: header in {"DOCK SEAL", "ACT KDT"},
    )
    if ddt_actual is not None and ddt_scheduled is not None:
        ddt_actual += ddt_scheduled + 1
    kat_actual = find_index(
        headers[kat_scheduled + 1 :] if kat_scheduled is not None else [],
        lambda header: header == "ACT KDT",
    )
    if kat_actual is not None and kat_scheduled is not None:
        kat_actual += kat_scheduled + 1

    return {
        "dock": find_index(headers, lambda header: header == "DOCK"),
        "loader": find_index(headers, lambda header: header in {"LOADER", "LOADERS"}),
        "driver": find_index(headers, lambda header: "DRIVER" in header),
        "truck": find_index(headers, lambda header: header == "TRUCK"),
        "flights": flight_indices[:3],
        "scheduledDdt": ddt_scheduled,
        "actualDdt": ddt_actual,
        "scheduledKat": kat_scheduled,
        "actualKat": kat_actual,
        "notes": find_index(headers, lambda header: header == "NOTES"),
    }


def cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return to_text(row[index])


def is_record_row(dock: str, row: tuple[Any, ...], columns: dict[str, Any]) -> bool:
    if norm(dock) == "DOCK":
        return False
    joined = " ".join(norm(value) for value in row)
    if "TOUR " in joined:
        return False
    flights = [cell(row, index) for index in columns["flights"]]
    has_dock = any(character.isdigit() for character in dock)
    has_staff_or_truck = any(
        [
            cell(row, columns["loader"]),
            cell(row, columns["driver"]),
            cell(row, columns["truck"]),
        ]
    )
    has_schedule = bool(cell(row, columns["scheduledDdt"]))
    if has_dock:
        return has_staff_or_truck or has_schedule or any(flights)
    return has_staff_or_truck and has_schedule and any(flights)


def extract_records(workbook: dict[str, Any]) -> list[dict[str, Any]]:
    workbook_path = workbook["path"]
    location = workbook["location"]
    tracker_page = f"{location} DDT Entry"
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if ws.title.upper() not in DAY_SHEETS and not ws.title.startswith("Touhy "):
            continue
        current_shift = "AM"
        current_date = sheet_date(ws, workbook["week_start"])
        headers = [norm(value) for value in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        columns = build_columns(headers)
        for row in ws.iter_rows(min_row=5, values_only=True):
            current_shift = detect_shift(list(row), current_shift)
            dock = cell(row, columns["dock"])
            if not is_record_row(dock, row, columns):
                continue
            flights = [
                {"flight": cell(row, index), "category": cell(row, index + 1)}
                for index in columns["flights"]
            ]
            if not any([dock, cell(row, columns["loader"]), cell(row, columns["truck"]), *[flight["flight"] for flight in flights]]):
                continue
            records.append(
                {
                    "id": f"history-{location.lower()}-{current_date}-{ws.title.lower().replace(' ', '-')}-{len(records) + 1}",
                    "location": location,
                    "trackerPage": tracker_page,
                    "date": current_date,
                    "shift": current_shift,
                    "dock": dock,
                    "loader": cell(row, columns["loader"]),
                    "driver": cell(row, columns["driver"]),
                    "truck": cell(row, columns["truck"]),
                    "flights": flights,
                    "scheduledDdt": cell(row, columns["scheduledDdt"]),
                    "actualDdt": cell(row, columns["actualDdt"]),
                    "scheduledKat": cell(row, columns["scheduledKat"]),
                    "actualKat": cell(row, columns["actualKat"]),
                    "delayReason": "",
                    "notes": cell(row, columns["notes"]),
                    "operationalComments": "",
                    "manager": "",
                    "supervisor": "",
                    "closedAt": f"{current_date}T23:59:00.000Z",
                }
            )
    return records


def build_snapshots() -> list[dict[str, Any]]:
    snapshots: list[dict[str, Any]] = []
    for workbook in WORKBOOKS:
        records = extract_records(workbook)
        by_date: dict[str, list[dict[str, Any]]] = {}
        for record in records:
            by_date.setdefault(record["date"], []).append(record)
        for day, day_records in sorted(by_date.items(), reverse=True):
            snapshots.append(
                {
                    "id": f"{workbook['location']}-{day}",
                    "location": workbook["location"],
                    "date": day,
                    "closedAt": f"{day}T23:59:00.000Z",
                    "records": day_records,
                }
            )
    return sorted(snapshots, key=lambda snapshot: snapshot["date"], reverse=True)


def main() -> None:
    snapshots = build_snapshots()
    SNAPSHOTS_OUT.write_text(json.dumps(snapshots, indent=2), encoding="utf-8")
    record_count = sum(len(snapshot["records"]) for snapshot in snapshots)
    print(f"wrote {len(snapshots)} snapshots and {record_count} records")


if __name__ == "__main__":
    main()
