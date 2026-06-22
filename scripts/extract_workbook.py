from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

WORKBOOK = Path("/Users/drewwelker/Documents/6-20-2026 DDT Tracker Touhy.xlsx")
ROOT = Path(__file__).resolve().parents[1]
SEED_OUT = ROOT / "src" / "data" / "seedRecords.json"
SCHEDULE_OUT = ROOT / "src" / "data" / "seedSchedules.json"
HISTORY_OUT = ROOT / "data" / "history" / "2026" / "06" / "20" / "touhy.json"


def to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def sheet_date(ws) -> str:
    value = ws["C1"].value
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return "2026-06-20"


def detect_shift(row_values: list[Any], current_shift: str) -> str:
    joined = " ".join(to_text(value).upper() for value in row_values)
    if "TOUR 1" in joined:
        return "Tour 1"
    if "TOUR 2" in joined:
        return "Tour 2"
    if "TOUR 3" in joined:
        return "Tour 3"
    return current_shift


def extract_records() -> list[dict[str, Any]]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if not ws.title.startswith("Touhy "):
            continue
        current_shift = "AM"
        current_date = sheet_date(ws)
        for row in ws.iter_rows(min_row=5, values_only=True):
            current_shift = detect_shift(list(row), current_shift)
            dock = to_text(row[0] if len(row) > 0 else "")
            loader = to_text(row[2] if len(row) > 2 else "")
            truck = to_text(row[4] if len(row) > 4 else "")
            flights = [
                {"flight": to_text(row[5] if len(row) > 5 else ""), "category": to_text(row[6] if len(row) > 6 else "")},
                {"flight": to_text(row[7] if len(row) > 7 else ""), "category": to_text(row[8] if len(row) > 8 else "")},
                {"flight": to_text(row[9] if len(row) > 9 else ""), "category": to_text(row[10] if len(row) > 10 else "")},
            ]
            if not any([dock, loader, truck, *[flight["flight"] for flight in flights]]):
                continue
            if not dock and "TOUR" in " ".join(to_text(v).upper() for v in row):
                continue
            records.append(
                {
                    "id": f"{ws.title.lower().replace(' ', '-')}-{len(records) + 1}",
                    "location": "Touhy",
                    "trackerPage": "Touhy DDT Entry",
                    "date": current_date,
                    "shift": current_shift,
                    "dock": dock,
                    "opsx": to_text(row[1] if len(row) > 1 else ""),
                    "loader": loader,
                    "driver": to_text(row[3] if len(row) > 3 else ""),
                    "truck": truck,
                    "flights": flights,
                    "scheduledDdt": to_text(row[11] if len(row) > 11 else ""),
                    "actualDdt": to_text(row[12] if len(row) > 12 else ""),
                    "scheduledKat": to_text(row[17] if len(row) > 17 else ""),
                    "actualKat": to_text(row[18] if len(row) > 18 else ""),
                    "delayReason": "",
                    "notes": to_text(row[20] if len(row) > 20 else ""),
                    "operationalComments": "",
                    "manager": "",
                    "supervisor": "",
                }
            )
    return records[:220]


def extract_schedule() -> list[dict[str, Any]]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    schedule: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if not ws.title.startswith("Touhy "):
            continue
        role_rows = [("Tower", ws["F1"].value), ("Dispatch", ws["F2"].value), ("Podium", ws["F3"].value)]
        for area, names in role_rows:
            if not to_text(names):
                continue
            schedule.append(
                {
                    "id": f"{ws.title.lower().replace(' ', '-')}-{area.lower()}",
                    "manager": to_text(names),
                    "supervisor": to_text(names),
                    "shift": "Daily coverage",
                    "startTime": "00:00",
                    "endTime": "23:59",
                    "area": area,
                    "date": sheet_date(ws),
                    "location": "Touhy",
                }
            )
    return schedule


def main() -> None:
    records = extract_records()
    schedules = extract_schedule()
    SEED_OUT.write_text(json.dumps(records, indent=2), encoding="utf-8")
    SCHEDULE_OUT.write_text(json.dumps(schedules, indent=2), encoding="utf-8")
    HISTORY_OUT.parent.mkdir(parents=True, exist_ok=True)
    HISTORY_OUT.write_text(json.dumps({"records": records[:40], "schedules": schedules}, indent=2), encoding="utf-8")
    print(f"wrote {len(records)} records and {len(schedules)} schedule rows")


if __name__ == "__main__":
    main()
